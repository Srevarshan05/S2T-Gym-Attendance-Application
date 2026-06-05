"""
app/utils/export_helpers.py
─────────────────────────────
Helper utilities for exporting data to CSV and Excel formats.
"""

import csv
import io
from typing import Any, Dict, Generator, List

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def generate_csv(data: List[Dict[str, Any]]) -> Generator[bytes, None, None]:
    """
    Convert a list of dictionaries into a memory-efficient CSV byte stream.
    Yields CSV rows as bytes encoded in UTF-8.
    
    If data is empty, yields an empty byte string.
    """
    if not data:
        yield b""
        return

    # Extract headers from the first dictionary's keys
    headers = list(data[0].keys())
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Yield header row
    writer.writerow(headers)
    yield output.getvalue().encode("utf-8")
    output.seek(0)
    output.truncate(0)
    
    # Yield data rows sequentially to save memory
    for row in data:
        writer.writerow([row.get(h) for h in headers])
        yield output.getvalue().encode("utf-8")
        output.seek(0)
        output.truncate(0)


def generate_excel(data: List[Dict[str, Any]]) -> io.BytesIO:
    """
    Convert a list of dictionaries into styled Excel file stream (openpyxl/pandas).
    Returns an in-memory BytesIO object representing the Excel file.
    
    Applies S2T palette colors to the headers:
      - Accent Background: #4ADE80 (Neon Green)
      - Primary Text: #111827 (Dark Gray, Bold)
    """
    df = pd.DataFrame(data)
    output = io.BytesIO()
    
    # Handle empty data edge case
    if df.empty:
        # Create an empty excel workbook
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Report")
        output.seek(0)
        return output

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Report")
        
        # Access the openpyxl workbook and worksheet objects
        worksheet = writer.sheets["Report"]
        
        # Style Definitions matching S2T design system
        header_fill = PatternFill(
            start_color="4ADE80",
            end_color="4ADE80",
            fill_type="solid"
        )
        header_font = Font(
            name="Segoe UI",
            size=11,
            bold=True,
            color="111827"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Style the header cells
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            
        # Give headers a generous row height
        worksheet.row_dimensions[1].height = 26
        
        # Auto-adjust column widths based on contents
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = cell.value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            # Add padding and set minimum column width
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output.seek(0)
    return output
